import cv2
import mediapipe as mp
import math
import time
import os
import threading
import pandas as pd
from fer import FER
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# ─── MediaPipe Pose (postura) ─────────────────────────────────────────────────
mp_pose = mp.solutions.pose
mp_draw = mp.solutions.drawing_utils
pose    = mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5)

# ─── MediaPipe FaceMesh (piscadas EAR) ───────────────────────────────────────
mp_face   = mp.solutions.face_mesh
face_mesh = mp_face.FaceMesh(
    max_num_faces=1, refine_landmarks=True,
    min_detection_confidence=0.5, min_tracking_confidence=0.5
)
RIGHT_EYE      = [33,  133, 160, 144, 158, 153]
LEFT_EYE       = [362, 263, 387, 373, 385, 380]
EAR_THRESH     = 0.20
FRAMES_FECHADO = 2

# ─── FER (emoção) ─────────────────────────────────────────────────────────────
fer_detector = FER(mtcnn=False)   # mtcnn=False → usa OpenCV, sem conflito com TF

EMOCAO_EMOJI = {
    "happy":    "FELIZ",
    "neutral":  "NEUTRO",
    "sad":      "TRISTE",
    "angry":    "RAIVA",
    "surprise": "SURPRESO",
    "fear":     "MEDO",
    "disgust":  "NOJO",
}

# ─── Câmera ───────────────────────────────────────────────────────────────────
cap = cv2.VideoCapture(0)
cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

# ─── Estado global ────────────────────────────────────────────────────────────
alpha  = 0.15
smooth_fwd = smooth_tilt = smooth_shrug = 0.0

bad_posture_start  = None
registros          = []
ultimo_registro    = time.time()
INTERVALO_REGISTRO = 10
id_contador        = 1
postura_atual      = "SEM DETECCAO"
score_atual        = 0.0
m_smooth           = {}

# Calibração
CALIBRANDO     = True
calib_frames   = []
CALIB_TOTAL    = 60
calib_baseline = {}

# Piscadas
piscadas_sessao     = 0
piscadas_intervalo  = 0
frames_olho_fechado = 0
olhos_fechados      = False
ear_atual           = 1.0

# Emoção — thread background
EMOCAO_ATUAL     = "analisando..."
EMOCAO_SCORES    = {}
EMOCAO_FRAME_BGR = None
_emocao_lock     = threading.Lock()
_emocao_running  = False


# ════════════════════════════════════════════════════════════════════════════════
#  POSTURA
# ════════════════════════════════════════════════════════════════════════════════

def extrair_metricas(landmarks, h, w):
    lm, VIS = landmarks, 0.4
    L_SH = lm[mp_pose.PoseLandmark.LEFT_SHOULDER.value]
    R_SH = lm[mp_pose.PoseLandmark.RIGHT_SHOULDER.value]
    L_EA = lm[mp_pose.PoseLandmark.LEFT_EAR.value]
    R_EA = lm[mp_pose.PoseLandmark.RIGHT_EAR.value]
    NOSE = lm[mp_pose.PoseLandmark.NOSE.value]
    m = {}
    offsets = []
    if L_SH.visibility > VIS and L_EA.visibility > VIS:
        offsets.append(abs(L_EA.x - L_SH.x))
    if R_SH.visibility > VIS and R_EA.visibility > VIS:
        offsets.append(abs(R_EA.x - R_SH.x))
    m["forward_head"]   = sum(offsets) / len(offsets) if offsets else 0.0
    m["head_tilt"]      = abs(L_EA.y - R_EA.y) if L_EA.visibility > VIS and R_EA.visibility > VIS else 0.0
    if L_SH.visibility > VIS and R_SH.visibility > VIS:
        m["shoulder_tilt"] = abs(L_SH.y - R_SH.y)
        sh_mid_y = (L_SH.y + R_SH.y) / 2
    else:
        m["shoulder_tilt"] = 0.0
        sh_mid_y = 0.5
    m["nose_drop"] = NOSE.y - sh_mid_y if NOSE.visibility > VIS else 0.0
    return m


def calcular_score_postura(m, baseline):
    fwd_base = max(baseline.get("forward_head",  0.01), 0.005)
    tlt_base = max(baseline.get("head_tilt",     0.01), 0.005)
    shl_base = max(baseline.get("shoulder_tilt", 0.01), 0.005)
    ndr_base =     baseline.get("nose_drop",     0.3)
    d_fwd = max(m["forward_head"]  - fwd_base * 1.5, 0) / fwd_base
    d_tlt = max(m["head_tilt"]     - tlt_base * 2.0, 0) / tlt_base
    d_shl = max(m["shoulder_tilt"] - shl_base * 2.0, 0) / shl_base
    d_ndr = max(ndr_base - m["nose_drop"] - 0.05,    0) * 10
    return min((d_fwd*50) + (d_tlt*20) + (d_shl*15) + (d_ndr*15), 100)


def classificar_postura(score):
    if score < 20:   return "BOA POSTURA",     (0, 200, 80)
    elif score < 50: return "POSTURA REGULAR", (0, 200, 220)
    else:            return "MA POSTURA",      (0, 60, 255)


def peso_postura(cl):
    return {"BOA POSTURA": 3, "POSTURA REGULAR": 2, "MA POSTURA": 1}.get(cl, 0)


# ════════════════════════════════════════════════════════════════════════════════
#  PISCADAS — EAR
# ════════════════════════════════════════════════════════════════════════════════

def dist(p1, p2):
    return math.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)

def calcular_ear(px, idx):
    p1,p4,p2,p6,p3,p5 = [px[i] for i in idx]
    return (dist(p2,p6) + dist(p3,p5)) / (2.0 * dist(p1,p4))

def processar_piscada(frame_rgb, h, w):
    res = face_mesh.process(frame_rgb)
    if not res.multi_face_landmarks:
        return 1.0, False, None
    lm = res.multi_face_landmarks[0].landmark
    px = [(int(p.x*w), int(p.y*h)) for p in lm]
    ear = (calcular_ear(px, RIGHT_EYE) + calcular_ear(px, LEFT_EYE)) / 2.0
    return ear, ear < EAR_THRESH, px


# ════════════════════════════════════════════════════════════════════════════════
#  EMOÇÃO — thread FER
# ════════════════════════════════════════════════════════════════════════════════

def _thread_emocao():
    global EMOCAO_ATUAL, EMOCAO_SCORES, _emocao_running
    while _emocao_running:
        with _emocao_lock:
            frame = EMOCAO_FRAME_BGR.copy() if EMOCAO_FRAME_BGR is not None else None
        if frame is None:
            time.sleep(0.05)
            continue
        try:
            resultado = fer_detector.detect_emotions(frame)
            if resultado:
                rosto    = max(resultado, key=lambda r: r["box"][2] * r["box"][3])
                scores   = rosto["emotions"]
                dominante = max(scores, key=scores.get)
                with _emocao_lock:
                    EMOCAO_ATUAL  = dominante
                    EMOCAO_SCORES = {k: round(v*100, 1) for k, v in scores.items()}
        except Exception:
            pass
        time.sleep(0.25)

def iniciar_thread_emocao():
    global _emocao_running
    _emocao_running = True
    t = threading.Thread(target=_thread_emocao, daemon=True)
    t.start()
    return t

def parar_thread_emocao():
    global _emocao_running
    _emocao_running = False


# ════════════════════════════════════════════════════════════════════════════════
#  OVERLAY
# ════════════════════════════════════════════════════════════════════════════════

def desenhar_overlay(frame, m_smooth, score, postura, cor, baseline,
                     pisc_sessao, pisc_intervalo, olhos_fech, ear, emocao, e_scores, h, w):
    # Postura
    cv2.putText(frame, postura, (50, 55), cv2.FONT_HERSHEY_SIMPLEX, 0.9, cor, 2)
    bx, by, bw, bh = 50, 75, 200, 16
    cv2.rectangle(frame, (bx, by), (bx+bw, by+bh), (50,50,50), -1)
    cv2.rectangle(frame, (bx, by), (bx+int(bw*score/100), by+bh), cor, -1)
    cv2.rectangle(frame, (bx, by), (bx+bw, by+bh), (160,160,160), 1)
    cv2.putText(frame, f"Score: {int(score)}/100", (bx, by-4), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (200,200,200), 1)

    # Piscadas
    c_olho = (0,60,255) if olhos_fech else (0,200,80)
    cv2.putText(frame, f"Olhos: {'FECHADOS' if olhos_fech else 'ABERTOS'}  EAR:{ear:.2f}",
                (50,120), cv2.FONT_HERSHEY_SIMPLEX, 0.65, c_olho, 2)
    cv2.putText(frame, f"Piscadas sessao: {pisc_sessao}",    (50,148), cv2.FONT_HERSHEY_SIMPLEX, 0.58, (200,200,200), 1)
    cv2.putText(frame, f"Piscadas intervalo: {pisc_intervalo}", (50,172), cv2.FONT_HERSHEY_SIMPLEX, 0.58, (200,200,200), 1)

    # Emoção
    label = EMOCAO_EMOJI.get(emocao, emocao.upper())
    cv2.putText(frame, f"Feicao: {label}", (50,200), cv2.FONT_HERSHEY_SIMPLEX, 0.72, (255,200,50), 2)

    # Barras de emoção (canto direito)
    if e_scores:
        ex, ey = w-230, 40
        cv2.rectangle(frame, (ex-5, ey-20), (w-10, ey+len(e_scores)*22+5), (30,30,30), -1)
        for i, (emo, val) in enumerate(sorted(e_scores.items(), key=lambda x: -x[1])):
            filled_w = int(180*val/100)
            c_b = (0,200,80) if emo == emocao else (100,100,100)
            cv2.rectangle(frame, (ex, ey+i*22), (ex+filled_w, ey+i*22+16), c_b, -1)
            cv2.putText(frame, f"{emo[:7]:<7} {val:5.1f}%",
                        (ex, ey+i*22+13), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (220,220,220), 1)

    # Métricas debug
    y0 = 232
    for k, v in m_smooth.items():
        cv2.putText(frame, f"{k}: {v:.4f}  (base:{baseline.get(k,0):.4f})",
                    (50, y0), cv2.FONT_HERSHEY_SIMPLEX, 0.38, (130,130,130), 1)
        y0 += 17


# ════════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ════════════════════════════════════════════════════════════════════════════════

def gerar_excel(registros, caminho):
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Monitoramento de Postura"
    COR_HDR="2563EB"; COR_BOA="D1FAE5"; COR_REG="FEF9C3"; COR_RUI="FEE2E2"
    thin  = Side(style="thin", color="CBD5E1")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["ID","Data","Horario","Piscadas","Postura","Peso","Score","Emocao"]
    col_widths = [8,14,12,12,24,10,10,14]
    for ci,(h,cw) in enumerate(zip(headers,col_widths),1):
        c=ws.cell(row=1,column=ci,value=h)
        c.font=Font(name="Arial",bold=True,color="FFFFFF",size=11)
        c.fill=PatternFill("solid",start_color=COR_HDR)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=borda
        ws.column_dimensions[get_column_letter(ci)].width=cw
    ws.row_dimensions[1].height=28

    total=len(registros)
    for i,reg in enumerate(registros,2):
        cl=reg["postura"]
        cor=COR_BOA if cl=="BOA POSTURA" else (COR_REG if cl=="POSTURA REGULAR" else COR_RUI)
        vals=[reg["id"],reg["data"],reg["horario"],reg["piscadas"],
              cl,reg["peso"],reg.get("score",0),reg.get("emocao","?")]
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=i,column=ci,value=val)
            c.font=Font(name="Arial",size=10)
            c.fill=PatternFill("solid",start_color=cor)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=borda
        ws.row_dimensions[i].height=20

    ws2=wb.create_sheet("Estatisticas")
    for ci,h in enumerate(["Classificacao","Quantidade","% do Total","Media Piscadas"],1):
        c=ws2.cell(row=1,column=ci,value=h)
        c.font=Font(name="Arial",bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",start_color=COR_HDR)
        c.alignment=Alignment(horizontal="center"); c.border=borda
    for row,(cat,cor_hex) in enumerate(zip(["BOA POSTURA","POSTURA REGULAR","MA POSTURA"],
                                            [COR_BOA,COR_REG,COR_RUI]),2):
        rc=[r for r in registros if r["postura"]==cat]
        qtd=len(rc)
        pct=round(qtd/total*100,1) if total else 0
        med=round(sum(r["piscadas"] for r in rc)/qtd,1) if qtd else 0
        for ci,val in enumerate([cat,qtd,pct,med],1):
            c=ws2.cell(row=row,column=ci,value=val)
            c.font=Font(name="Arial",size=10)
            c.fill=PatternFill("solid",start_color=cor_hex)
            c.alignment=Alignment(horizontal="center"); c.border=borda
    for col,cw in zip(["A","B","C","D"],[22,12,12,18]):
        ws2.column_dimensions[col].width=cw
    ws.freeze_panes="A2"
    wb.save(caminho)
    print(f"✅ Excel salvo em: {caminho}")


# ════════════════════════════════════════════════════════════════════════════════
#  LOOP PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════════

NOME_JANELA = "Monitor Postura + Piscadas + Emocao"
cv2.namedWindow(NOME_JANELA, cv2.WINDOW_NORMAL)
print("Sistema iniciado! R = recalibrar | Q = sair")
print("FER carregado — deteccao de emocoes ATIVA")

thread_emocao = iniciar_thread_emocao()

try:
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        agora     = time.time()
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, _   = frame.shape
        results   = pose.process(frame_rgb)

        postura_atual = "SEM DETECCAO"
        cor           = (128, 128, 128)

        if results.pose_landmarks:
            mp_draw.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS)
            metricas = extrair_metricas(results.pose_landmarks.landmark, h, w)

            if CALIBRANDO:
                calib_frames.append(metricas)
                prog = len(calib_frames)
                cv2.putText(frame, f"CALIBRANDO... sente-se reto! {prog}/{CALIB_TOTAL}",
                            (50,60), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0,255,255), 2)
                if prog >= CALIB_TOTAL:
                    for k in metricas:
                        calib_baseline[k] = sum(f[k] for f in calib_frames) / len(calib_frames)
                    CALIBRANDO = False
                    print("✅ Calibração concluída:", calib_baseline)
                cv2.imshow(NOME_JANELA, frame)
                if cv2.waitKey(10) & 0xFF == ord("q"):
                    break
                continue

            smooth_fwd   = alpha*metricas["forward_head"]  + (1-alpha)*smooth_fwd
            smooth_tilt  = alpha*metricas["head_tilt"]     + (1-alpha)*smooth_tilt
            smooth_shrug = alpha*metricas["shoulder_tilt"] + (1-alpha)*smooth_shrug
            smooth_nose  = alpha*metricas["nose_drop"]     + (1-alpha)*calib_baseline.get("nose_drop",0.3)

            m_smooth = {"forward_head": smooth_fwd, "head_tilt": smooth_tilt,
                        "shoulder_tilt": smooth_shrug, "nose_drop": smooth_nose}

            score_atual   = calcular_score_postura(m_smooth, calib_baseline)
            postura_atual, cor = classificar_postura(score_atual)

            bad_posture_start = agora if postura_atual == "MA POSTURA" else None
            if bad_posture_start and postura_atual == "MA POSTURA":
                bad_posture_start = bad_posture_start or agora
                elapsed = int(agora - bad_posture_start)
                cv2.putText(frame, f"Tempo ruim: {elapsed}s",
                            (50, h-50), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0,60,255), 2)

        # Emoção — envia frame para thread e lê resultado
        with _emocao_lock:
            EMOCAO_FRAME_BGR = frame.copy()
        with _emocao_lock:
            emocao_agora = EMOCAO_ATUAL
            scores_agora = dict(EMOCAO_SCORES)

        # Piscadas
        ear_atual, fechado_agora, lm_px = processar_piscada(frame_rgb, h, w)
        if lm_px:
            for idx in RIGHT_EYE + LEFT_EYE:
                cv2.circle(frame, lm_px[idx], 2, (0,255,255), -1)
        if fechado_agora:
            frames_olho_fechado += 1
        else:
            if frames_olho_fechado >= FRAMES_FECHADO:
                piscadas_sessao    += 1
                piscadas_intervalo += 1
            frames_olho_fechado = 0
        olhos_fechados = fechado_agora

        # Overlay
        if not CALIBRANDO and results.pose_landmarks:
            desenhar_overlay(frame, m_smooth, score_atual, postura_atual, cor,
                             calib_baseline, piscadas_sessao, piscadas_intervalo,
                             olhos_fechados, ear_atual, emocao_agora, scores_agora, h, w)

        # HUD inferior
        tempo_prox = INTERVALO_REGISTRO - int(agora - ultimo_registro)
        cv2.putText(frame, f"Prox. registro: {tempo_prox}s | Registros: {len(registros)}",
                    (50, h-20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (170,170,170), 1)
        cv2.putText(frame, "R=recalibrar | Q=sair",
                    (w-240, h-20), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (170,170,170), 1)

        # Registro periódico
        if (not CALIBRANDO and agora-ultimo_registro >= INTERVALO_REGISTRO
                and postura_atual != "SEM DETECCAO"):
            now_dt = datetime.now()
            registros.append({
                "id": id_contador, "data": now_dt.strftime("%d/%m/%Y"),
                "horario": now_dt.strftime("%H:%M:%S"), "piscadas": piscadas_intervalo,
                "postura": postura_atual, "peso": peso_postura(postura_atual),
                "score": round(score_atual,1), "emocao": emocao_agora,
            })
            print(f"[{now_dt.strftime('%H:%M:%S')}] #{id_contador}: {postura_atual} | "
                  f"score={score_atual:.1f} | piscadas={piscadas_intervalo} | emocao={emocao_agora}")
            id_contador       += 1
            ultimo_registro    = agora
            piscadas_intervalo = 0

        cv2.imshow(NOME_JANELA, frame)

        key = cv2.waitKey(10) & 0xFF
        if key == ord("q"):
            parar_thread_emocao()
            if not CALIBRANDO and postura_atual != "SEM DETECCAO":
                now_dt = datetime.now()
                registros.append({
                    "id": id_contador, "data": now_dt.strftime("%d/%m/%Y"),
                    "horario": now_dt.strftime("%H:%M:%S"), "piscadas": piscadas_intervalo,
                    "postura": postura_atual, "peso": peso_postura(postura_atual),
                    "score": round(score_atual,1), "emocao": emocao_agora,
                })
            break
        elif key == ord("r") and not CALIBRANDO:
            CALIBRANDO = True; calib_frames = []; calib_baseline = {}
            smooth_fwd = smooth_tilt = smooth_shrug = 0.0
            print("🔄 Recalibrando...")

except KeyboardInterrupt:
    parar_thread_emocao()
    print("\n⚠️  Interrompido (Ctrl+C). Salvando...")
    if not CALIBRANDO and postura_atual != "SEM DETECCAO":
        now_dt = datetime.now()
        registros.append({
            "id": id_contador, "data": now_dt.strftime("%d/%m/%Y"),
            "horario": now_dt.strftime("%H:%M:%S"), "piscadas": piscadas_intervalo,
            "postura": postura_atual, "peso": peso_postura(postura_atual),
            "score": round(score_atual,1), "emocao": EMOCAO_ATUAL,
        })
finally:
    cap.release()
    cv2.destroyAllWindows()

# ════════════════════════════════════════════════════════════════════════════════
#  EXPORTAÇÃO FINAL
# ════════════════════════════════════════════════════════════════════════════════

if registros:
    pasta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
    os.makedirs(pasta, exist_ok=True)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")

    df = pd.DataFrame(registros)
    caminho_csv = os.path.join(pasta, f"postura_{ts}.csv")
    df.to_csv(caminho_csv, index=False, encoding="utf-8-sig")
    print(f"✅ CSV salvo em: {caminho_csv}")

    caminho_txt = os.path.join(pasta, f"postura_{ts}.txt")
    total = len(df)
    with open(caminho_txt, "w", encoding="utf-8") as f:
        f.write("="*55+"\n  RELATÓRIO — POSTURA + PISCADAS + EMOÇÃO\n"+"="*55+"\n\n")
        f.write(f"Gerado em          : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write(f"Total de registros : {total}\n")
        f.write(f"Score medio        : {df['score'].mean():.1f}/100\n")
        f.write(f"Total de piscadas  : {piscadas_sessao}\n\n")
        f.write("── Distribuicao de emocoes ──\n")
        emos = [r.get("emocao","?") for r in registros]
        for emo in sorted(set(emos)):
            qtd = emos.count(emo)
            f.write(f"  {emo:<12} {qtd:>3} registros ({qtd/total*100:.1f}%)\n")
        f.write("\n── Distribuicao de postura ──\n")
        for cat, qtd in df["postura"].value_counts().items():
            f.write(f"  {cat:<22} {qtd:>3} registros ({qtd/total*100:.1f}%)\n")
        f.write("\n── Detalhes ──\n")
        for _, row in df.iterrows():
            f.write(f"  [{row['horario']}] #{int(row['id']):03d}  "
                    f"{row['postura']:<22} score={row['score']:.1f}  "
                    f"piscadas={int(row['piscadas'])}  emocao={row.get('emocao','?')}\n")
    print(f"✅ TXT salvo em: {caminho_txt}")

    caminho_xlsx = os.path.join(pasta, f"postura_{ts}.xlsx")
    gerar_excel(registros, caminho_xlsx)

    try:
        if os.name == "nt":
            os.startfile(caminho_xlsx)
        else:
            cmd = "open" if os.uname().sysname == "Darwin" else "xdg-open"
            os.system(f'{cmd} "{caminho_xlsx}"')
    except Exception as e:
        print(f"Nao foi possivel abrir automaticamente: {e}")

else:
    print(f"\nNenhum registro coletado. O programa registra a cada {INTERVALO_REGISTRO}s.")